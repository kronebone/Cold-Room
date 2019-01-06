import datetime
from openpyxl import Workbook, load_workbook
from time import sleep
from sense_hat import SenseHat
from twilio.rest import Client
from collections import deque


class ColdRoom:
    def __init__(self):
        self.sense = SenseHat()
        self.temps_24_hrs = deque([], maxlen=24)
        self.temps_7_days = deque([], maxlen=7)
        self.temps_30_days = deque([], maxlen=30)
        self.log_created = False

    def log_temps(self, log_of_temps, temperature):
        # adds the taken temperature to the right log
        log_of_temps.append(temperature)

    def write_temps(self, filename):
        if not self.log_created:
            temps_wb = Workbook()
            temps_wb.save(filename)
            self.log_created = True
        else:
            temps_wb = load_workbook(filename)
            sheet = temps_wb.active
            sheet['A1'] = str(self.show_24_hrs())
            sheet['A2'] = str(self.show_7_days())
            sheet['A3'] = str(self.show_30_days())
            temps_wb.save(filename)

    def show_avg_temps(self, taken_temps):
        # averages the temps in a given list
        if len(taken_temps) == 0:
            return "No temps recorded yet."

        elif taken_temps > 0:
            total = sum(taken_temps)
            return int(total / len(taken_temps))

    def show_24_hrs(self):
        # easier way to get avg temps
        return self.show_avg_temps(taken_temps=self.temps_24_hrs)

    def show_7_days(self):
        # easier way to get avg temps
        return self.show_avg_temps(taken_temps=self.temps_7_days)

    def show_30_days(self):
        # easier way to get avg temps
        return self.show_avg_temps(taken_temps=self.temps_30_days)

    def adjusted_temp(self):
        # gets the temperature from the 3 sense hat sensors
        thermometer_temp = self.sense.get_temperature() * 9 / 5 + 32
        pressure_temp = self.sense.get_temperature_from_pressure() * 9 / 5 + 32
        humidity_temp = self.sense.get_temperature_from_humidity() * 9 / 5 + 32
        adjusted_temp = ((thermometer_temp + pressure_temp + humidity_temp) / 3)
        return adjusted_temp - adjusted_temp * .25  # used to help offset the heat the raspberry pi emits

    def pixel_setter(self, pixel_color, pixel_list):
        # used to create a list of rgb values for the led array
        for i in range(64):
            pixel_list.append(pixel_color)

    def show_temp_number(self, temp_color):
        # scrolls the temperature number
        temperature = int(self.adjusted_temp())
        self.sense.show_message(text_string=str(temperature),
                                text_colour=[0, 0, 0],
                                back_colour=temp_color,
                                scroll_speed=.4)

    def text_alert(self):
        # sends the alert text using twilio
        account_sid = ''  # twilio sid
        auth_token = ''  # twilio token
        client = Client(account_sid, auth_token)
        # phone numbers in contact list receive a text alert
        contact_list = []
        for contact in contact_list:
            client.messages.create(
                to=contact,
                from_='',  # twilio phone number
                body="Server room is over 80F")
        text_sent_time_string = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        text_sent_time = datetime.datetime.strptime(text_sent_time_string, '%Y-%m-%d %H:%M')
        return text_sent_time

    def led_temp(self):
        # reads the temp and updates the led array, calls the alert text function if temps are too high
        full_array = []
        cold = [2, 71, 254]  # blue
        warm = [255, 246, 0]  # yellow
        hot = [255, 8, 0]  # red
        done = False
        current_reading = self.adjusted_temp()

        if current_reading < 75:
            self.show_temp_number(cold)
            self.pixel_setter(cold, full_array)
            self.sense.set_pixels(full_array)

        elif 75 <= current_reading < 80:
            self.show_temp_number(warm)
            self.pixel_setter(warm, full_array)
            self.sense.set_pixels(full_array)

        elif current_reading >= 80:
            self.show_temp_number(hot)
            self.pixel_setter(hot, full_array)
            time_sent_alert = self.text_alert()
            wait_time = datetime.timedelta(minutes=45)
            self.sense.set_pixels(full_array)
            while not done:
                # delays another text from being sent after sending the first
                # if the room is still too hot it will send another text after waiting
                current_time_string = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                current_time = datetime.datetime.strptime(current_time_string, '%Y-%m-%d %H:%M')
                if current_time < time_sent_alert + wait_time:
                    self.show_temp_number(hot)
                    sleep(60)
                elif current_time >= time_sent_alert + wait_time:
                    done = True
