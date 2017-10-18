import datetime
from openpyxl import Workbook, load_workbook
from time import sleep
from sense_hat import SenseHat
from twilio.rest import Client


class Cold_room:
    def __init__(self):
        self.sense = SenseHat()
        self.temps_24_hrs = []
        self.temps_7_days = []
        self.temps_30_days = []
        self.log_created = False

    def list_length_checker(self):
        # keeps the temperature lists a certain length
        if len(self.temps_24_hrs) <= 48:
            pass
        elif len(self.temps_24_hrs) == 49:
            del self.temps_24_hrs[0]

        if len(self.temps_7_days) <= 7:
            pass
        elif len(self.temps_7_days) == 8:
            del self.temps_7_days[0]

        if len(self.temps_30_days) <= 30:
            pass
        elif len(self.temps_30_days) == 31:
            del self.temps_30_days[0]

    def log_temps(self, log_of_temps, temperature):
        # adds the taken temperature to the right log
        log_of_temps.append(temperature)
        self.list_length_checker()

    def write_temps(self, filename):
        if self.log_created == False:
            temps_wb = Workbook()
            temps_wb.save(filename)
            self.log_created = True
        else:
            temps_wb = load_workbook(filename)
            sheet = temps_wb.active
            sheet['A1'] = str(self.show_24_hrs())
            sheet['A2'] = str(self.show_7_days())
            sheet['A3'] = str(self.show_30_days())
            temps_wb.save(filename)



    def show_avg_temps(self, taken_temps):
        # averages the temps in a given list
        if taken_temps == []:
            return "No temps recorded yet."

        elif taken_temps != []:
            total = 0
            for t in taken_temps:
                total += t
            return int(total / int(len(taken_temps)))


    def show_24_hrs(self):
        # easier way to get avg temps
        return self.show_avg_temps(taken_temps=self.temps_24_hrs)

    def show_7_days(self):
        # easier way to get avg temps
        return self.show_avg_temps(taken_temps=self.temps_7_days)

    def show_30_days(self):
        # easier way to get avg temps
        return  self.show_avg_temps(taken_temps=self.temps_30_days)


    def adjusted_temp(self):
        # gets the temperature from the 3 sense hat sensors converts to F and adjusts for the heat of the pi
        thermometer_temp = self.sense.get_temperature() * 9 / 5 + 32
        pressure_temp = self.sense.get_temperature_from_pressure() * 9 / 5 + 32
        humidity_temp = self.sense.get_temperature_from_humidity() * 9 / 5 + 32
        adjusted_temp = ((thermometer_temp + pressure_temp + humidity_temp) / 3)
        return adjusted_temp - adjusted_temp * .25

    def pixel_setter(self, pixel_color, pixel_list):
        # used to create a list of rgb values for the led array
        for i in range(64):
            pixel_list.append(pixel_color)

    def show_temp_number(self, temp_color):
        # scrolls the temperature number
        temperature = int(self.adjusted_temp())
        self.sense.show_message(text_string=str(temperature), text_colour=[0, 0, 0], back_colour=temp_color, scroll_speed=.4)

    def text_alert(self):
        # sends the alert text using twilio
        account_sid = ''
        auth_token = ''
        client = Client(account_sid, auth_token)
        employee1_phone = ''
        employee2_phone = ''
        employee3_phone = ''
        contact_list = [employee1_phone, employee2_phone, employee3_phone]
        for contact in contact_list:
            client.messages.create(
                to=contact,
                from_='',
                body="Server room is over 80F")
        text_sent_time_string = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        text_sent_time = datetime.datetime.strptime(text_sent_time_string, '%Y-%m-%d %H:%M')
        return text_sent_time

    def led_temp(self):
        # reads the temp and updates the led array, calls the alert text function if temps are too high
        full_array = []
        cold = [2, 71, 254]  # blue
        warm = [255, 246, 0]  # orange
        hot = [255, 8, 0]  # red
        done = False
        current_reading = self.adjusted_temp()

        if current_reading < 75:
            self.show_temp_number(cold)
            self.pixel_setter(cold, full_array)
            self.sense.set_pixels(full_array)

        elif 75 <= current_reading < 80:
            self.show_temp_number(warm)
            self.pixel_setter(warm, full_array)
            self.sense.set_pixels(full_array)

        elif current_reading >= 80:
            self.show_temp_number(hot)
            self.pixel_setter(hot, full_array)
            time_sent_alert = self.text_alert()
            thirty_min_wait = datetime.timedelta(minutes=45)
            self.sense.set_pixels(full_array)
            while not done:
                # delays everything 45 minutes from the time an alert text was sent to prevent too many texts being sent out
                current_time_string = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                current_time = datetime.datetime.strptime(current_time_string, '%Y-%m-%d %H:%M')
                if current_time < time_sent_alert + thirty_min_wait:
                    self.show_temp_number(hot)
                    sleep(60)
                elif current_time >= time_sent_alert + thirty_min_wait:
                    done = True
